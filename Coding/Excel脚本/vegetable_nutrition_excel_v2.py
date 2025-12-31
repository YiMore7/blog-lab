# 在脚本开头添加
print("===== 脚本开始执行 =====")
import pandas as pd
import numpy as np
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

print(f"Python版本: {sys.version}")
print(f"当前工作目录: {os.getcwd()}")

# 创建蔬菜数据
vegetable_data = {
    '蔬菜科属': ['藜科', '十字花科', '伞形科', '茄科', '十字花科', '葫芦科', '旋花科', '百合科', '十字花科', '百合科'],
    '蔬菜代表': ['菠菜', '西兰花', '胡萝卜', '番茄', '甘蓝', '南瓜', '红薯', '芦笋', '菜花', '洋葱'],
    '主要营养成分': ['叶酸、铁、维生素K', '维生素C、叶酸、钾', '胡萝卜素、维生素A', '番茄红素、维生素C', '维生素K、维生素C', '胡萝卜素、维生素A', '胡萝卜素、维生素C', '叶酸、维生素K', '维生素C、叶酸', '槲皮素、硫化物'],
    '营养亮点': ['补血佳品', '抗癌明星', '护眼高手', '美容圣品', '骨骼守护者', '肠道清道夫', '血糖稳定剂', '排毒先锋', '免疫增强剂', '心血管保护神']
}

# 创建DataFrame
df = pd.DataFrame(vegetable_data)

# 按照主要营养成分排序（这里简化处理，按字符串顺序排列）
# 实际应用中可以根据具体营养成分进行更精确的排序
# df = df.sort_values(by='主要营养成分')

# 创建Excel工作簿
wb = Workbook()
ws = wb.active
ws.title = "蔬菜营养排行"

# 添加标题
ws.merge_cells('A1:D1')
ws['A1'] = '蔬菜营养成分排行表'
ws['A1'].font = Font(size=16, bold=True)
ws['A1'].alignment = Alignment(horizontal='center')

# 添加数据来源和说明
ws.merge_cells('A2:D2')
ws['A2'] = '数据来源：基于科学研究的蔬菜营养成分分析'
ws['A2'].alignment = Alignment(horizontal='center')

# 添加列标题
headers = ['蔬菜科属', '蔬菜代表', '主要营养成分', '营养亮点']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')

# 添加数据
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 4):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.value = value
        cell.alignment = Alignment(horizontal='center' if c_idx != 3 else 'left')
        
# 设置列宽
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 20

# 添加边框
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in ws.iter_rows(min_row=3, max_row=13, min_col=1, max_col=4):
    for cell in row:
        cell.border = thin_border

# 添加筛选
ws.auto_filter.ref = "A3:D13"

# 冻结窗格
ws.freeze_panes = 'A4'

# 保存Excel文件
try:
    file_path = os.path.join(os.getcwd(), '蔬菜营养排行表_更新版.xlsx')
    print(f"尝试保存Excel文件到: {file_path}")
    wb.save(file_path)
    print(f"Excel文件已成功创建！保存在: {file_path}")
except Exception as e:
    print(f"创建Excel文件时出错: {e}")

# 在脚本结尾添加
print("===== 脚本执行完毕 =====")